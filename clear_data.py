import os
import openpyxl
import win32com.client
import config

SHEETS_TO_CLEAR = [
    config.SHEET_PASSENGER,
    config.SHEET_STUDENT_DETAILS,
    config.SHEET_STAFF_DETAILS,
    config.SHEET_ERROR,
]


def _close_if_open(path: str):
    try:
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return
    target = os.path.abspath(path).lower()
    for wb in excel.Workbooks:
        if wb.FullName.lower() == target:
            wb.Save()
            wb.Close(False)
            print("Closed workbook before clearing.")
            return


_close_if_open(config.EXCEL_FILE)

wb = openpyxl.load_workbook(config.EXCEL_FILE)

for name in SHEETS_TO_CLEAR:
    if name not in wb.sheetnames:
        print(f"  Skipped (not found): {name}")
        continue
    ws = wb[name]
    rows_cleared = 0
    for row in ws.iter_rows(min_row=2):
        if any(cell.value is not None for cell in row):
            for cell in row:
                cell.value = None
            rows_cleared += 1
    print(f"  Cleared {rows_cleared} row(s): {name}")

wb.save(config.EXCEL_FILE)
print("Done.")
