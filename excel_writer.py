import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import config


def _as_number(val):
    """Return val as int if it's a digit-only string, otherwise return as-is."""
    if isinstance(val, str) and val.isdigit():
        return int(val)
    return val


def ensure_headers(ws):
    """Write header row, adding any missing columns to existing files."""
    for col_idx, header in enumerate(config.HEADERS, start=1):
        if ws.cell(row=1, column=col_idx).value != header:
            ws.cell(row=1, column=col_idx).value = header


def get_next_row(ws) -> int:
    """Return the first empty row in column A (minimum 2 to leave room for headers)."""
    max_row = ws.max_row
    if max_row < 2:
        return 2
    # Walk back from max_row to find the actual last filled row
    for row in range(max_row, 1, -1):
        if ws.cell(row=row, column=config.COL_ENTRY_ID).value is not None:
            return row + 1
    return 2


def write_row(ws, row_num: int, row_data: dict):
    """Write a single passenger dict into the given row."""
    col_map = {
        "EntryID":               config.COL_ENTRY_ID,
        "PNR":                   config.COL_PNR,
        "PassengerName":         config.COL_PASSENGER_NAME,
        "AeroplanNumber":        config.COL_AEROPLAN,
        "FirstDepartureAirport": config.COL_FIRST_DEP,
        "OutboundSegments":      config.COL_OUTBOUND_SEG,
        "ReturnSegments":        config.COL_RETURN_SEG,
        "MontrealArrivalTime":   config.COL_MTL_ARRIVAL,
        "MontrealDepartureTime": config.COL_MTL_DEPARTURE,
        "FlightPassProduct":     config.COL_FP_PRODUCT,
        "CreditsPerPassenger":   config.COL_CREDITS_PER_PAX,
        "Cost":                  config.COL_COST,
        "Type":                  config.COL_TYPE,
    }
    for key, col in col_map.items():
        val = row_data.get(key, "")
        if key == "AeroplanNumber":
            val = _as_number(val)
        ws.cell(row=row_num, column=col).value = val


def get_all_entry_ids(ws) -> set:
    """Return a set of all EntryID values from column A (rows 2+)."""
    ids = set()
    for row in ws.iter_rows(min_row=2, min_col=config.COL_ENTRY_ID,
                             max_col=config.COL_ENTRY_ID, values_only=True):
        val = row[0]
        if val is not None:
            ids.add(str(val))
    return ids


def _col_letter(col_idx: int) -> str:
    """Convert 1-based column index to Excel letter (A, B, … Z, AA …)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _auto_fit_column(ws, col_idx: int, extra: int = 2):
    """Set column width based on the longest value in that column."""
    col_letter = _col_letter(col_idx)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            # Use the longest line within a wrapped cell
            lines = str(cell.value).split("\n")
            cell_max = max(len(line) for line in lines)
            max_len = max(max_len, cell_max)
    ws.column_dimensions[col_letter].width = max_len + extra


def format_passenger_sheet(ws):
    """
    Apply formatting to the PassengerData sheet:
    - Column widths
    - Wrap text on F and G
    - Bold grey header row
    - Alternating row banding
    """
    last_row = ws.max_row
    if last_row < 1:
        return

    # --- turn off wrap text everywhere first ---
    for row in ws.iter_rows(min_row=1, max_row=last_row,
                             min_col=1, max_col=len(config.HEADERS)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    # --- column A fixed width ---
    ws.column_dimensions[_col_letter(config.COL_ENTRY_ID)].width = 7

    # --- auto-fit B through M (first pass, before wrap) ---
    for col_idx in range(config.COL_PNR, len(config.HEADERS) + 1):
        _auto_fit_column(ws, col_idx)

    # --- enable wrap text on F and G ---
    for row_num in range(1, last_row + 1):
        ws.cell(row=row_num, column=config.COL_OUTBOUND_SEG).alignment = \
            Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row_num, column=config.COL_RETURN_SEG).alignment = \
            Alignment(wrap_text=True, vertical="top")

    # --- re-auto-fit B through M (second pass, after wrap) ---
    for col_idx in range(config.COL_PNR, len(config.HEADERS) + 1):
        _auto_fit_column(ws, col_idx)

    # --- header row formatting ---
    header_fill = PatternFill(fill_type="solid", fgColor=config.COLOR_HEADER)
    for col_idx in range(1, len(config.HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # --- row banding (data rows 2+) ---
    band_fill = PatternFill(fill_type="solid", fgColor=config.COLOR_ROW_BAND)
    no_fill   = PatternFill(fill_type=None)

    for row_num in range(2, last_row + 1):
        # VBA: i Mod 2 = 1 → rows 3, 5, 7 … get the band
        fill = band_fill if row_num % 2 == 1 else no_fill
        for col_idx in range(1, len(config.HEADERS) + 1):
            ws.cell(row=row_num, column=col_idx).fill = fill

    # --- row heights ---
    for row_num in range(1, last_row + 1):
        outbound = ws.cell(row=row_num, column=config.COL_OUTBOUND_SEG).value or ""
        return_   = ws.cell(row=row_num, column=config.COL_RETURN_SEG).value or ""
        line_count = max(
            str(outbound).count("\n") + 1,
            str(return_).count("\n") + 1,
            1
        )
        if line_count > 1:
            ws.row_dimensions[row_num].height = line_count * config.DEFAULT_ROW_HEIGHT
        else:
            ws.row_dimensions[row_num].height = config.DEFAULT_ROW_HEIGHT


DETAILS_HEADERS = [
    "Name Combined", "Preferred Name", "Email", "School/Inst",
    "Position/title", "Conference Arrival", "AeroplanNumber",
    "PNR", "Type", "FirstDepartureAirport", "OutboundSegments",
    "ReturnSegments", "MontrealArrivalTime", "MontrealDepartureTime",
    "FlightPassProduct", "CreditsPerPassenger", "Cost", "EmailStatus",
]

ERROR_HEADERS = ["EntryID", "PNR", "PassengerName", "AeroplanNumber", "Reason"]


def _get_next_row_any(ws) -> int:
    """First empty row in column A of any sheet (minimum 2)."""
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=1).value is not None:
            return row + 1
    return 2


def ensure_details_sheet(wb, sheet_name: str):
    """Create details sheet if it doesn't exist, and ensure row 1 has all headers."""
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    header_fill = PatternFill(fill_type="solid", fgColor=config.COLOR_HEADER)
    for col_idx, header in enumerate(DETAILS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
            cell.font  = Font(bold=True)
            cell.fill  = header_fill
    return ws


def ensure_error_sheet(wb):
    """Create Error sheet if it doesn't exist, and ensure row 1 has all headers."""
    if config.SHEET_ERROR not in wb.sheetnames:
        wb.create_sheet(config.SHEET_ERROR)
    ws = wb[config.SHEET_ERROR]
    err_fill = PatternFill(fill_type="solid", fgColor="FF9999")
    for col_idx, header in enumerate(ERROR_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
            cell.font  = Font(bold=True)
            cell.fill  = err_fill
    return ws


def write_details_row(ws, row_num: int, passenger_data: dict, reg_data: dict):
    """Write merged registration + flight data to a details sheet row."""
    values = [
        reg_data.get("Name Combined", ""),
        reg_data.get("Preferred Name", ""),
        reg_data.get("Email", ""),
        reg_data.get("School/Inst", ""),
        reg_data.get("Position/title", ""),
        reg_data.get("Conference Arrival", ""),
        _as_number(passenger_data.get("AeroplanNumber", "")),
        passenger_data.get("PNR", ""),
        passenger_data.get("Type", ""),
        passenger_data.get("FirstDepartureAirport", ""),
        passenger_data.get("OutboundSegments", ""),
        passenger_data.get("ReturnSegments", ""),
        passenger_data.get("MontrealArrivalTime", ""),
        passenger_data.get("MontrealDepartureTime", ""),
        passenger_data.get("FlightPassProduct", ""),
        passenger_data.get("CreditsPerPassenger", ""),
        passenger_data.get("Cost", ""),
        "",  # EmailStatus — blank until previewed
    ]
    for col_idx, val in enumerate(values, start=1):
        ws.cell(row=row_num, column=col_idx).value = val


SKIP_HEADERS = ["EntryID", "Subject", "Note"]


def ensure_skip_sheet(wb):
    """Create the Do Not Import sheet if needed and ensure headers are present."""
    if config.SHEET_SKIP not in wb.sheetnames:
        wb.create_sheet(config.SHEET_SKIP)
    ws = wb[config.SHEET_SKIP]
    skip_fill = PatternFill(fill_type="solid", fgColor="FFD700")
    for col_idx, header in enumerate(SKIP_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
            cell.font  = Font(bold=True)
            cell.fill  = skip_fill
    return ws


def get_skip_ids(wb) -> set:
    """Return the set of EntryIDs from the Do Not Import sheet."""
    if config.SHEET_SKIP not in wb.sheetnames:
        return set()
    ws = wb[config.SHEET_SKIP]
    ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    return ids


def write_skip_row(ws, row_num: int, entry_id: str, subject: str, note: str = ""):
    """Write one entry to the Do Not Import sheet."""
    ws.cell(row=row_num, column=1).value = entry_id
    ws.cell(row=row_num, column=2).value = subject
    ws.cell(row=row_num, column=3).value = note


def write_error_row(ws, row_num: int, passenger_data: dict, reason: str):
    """Write an unmatched passenger to the Error sheet."""
    for col_idx, val in enumerate([
        passenger_data.get("EntryID", ""),
        passenger_data.get("PNR", ""),
        passenger_data.get("PassengerName", ""),
        _as_number(passenger_data.get("AeroplanNumber", "")),
        reason,
    ], start=1):
        ws.cell(row=row_num, column=col_idx).value = val


def log_debug(wb, entry_id: str, subject: str, error: str, row_num: int):
    """Append an error record to the Debug sheet."""
    if config.SHEET_DEBUG not in wb.sheetnames:
        ws_debug = wb.create_sheet(config.SHEET_DEBUG)
        ws_debug.append(["EntryID", "Subject", "Error", "RowNum"])
    else:
        ws_debug = wb[config.SHEET_DEBUG]
    ws_debug.append([entry_id, subject, error, row_num])
