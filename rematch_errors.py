"""
Re-attempt Aeroplan matching for every PassengerData row with MatchStatus = 'Error'.

When to use:
  Someone's Aeroplan number has been added (or corrected) in the Student or Staff
  registration sheet since the initial import.  Run this to pick them up without
  re-importing all emails.

What it does:
  1. Reads PassengerData rows where MatchStatus = 'Error'
  2. Tries lookup_aeroplan() against the current Student/Staff sheets
  3. If found: writes a new row in the appropriate Plane Details sheet,
               sets MatchStatus = 'Staff' or 'Student', removes the Error sheet row
  4. If still not found: leaves MatchStatus = 'Error' and reports it
"""
import os
import sys
import traceback
import openpyxl
import win32com.client

import config
from match_passenger import lookup_aeroplan, details_sheet_for
from excel_writer import (
    ensure_details_sheet, write_details_row, _get_next_row_any,
)


def _read_passenger_row_openpyxl(ws, row_num: int) -> dict:
    return {
        "EntryID":               ws.cell(row=row_num, column=config.COL_ENTRY_ID).value,
        "PNR":                   ws.cell(row=row_num, column=config.COL_PNR).value,
        "PassengerName":         ws.cell(row=row_num, column=config.COL_PASSENGER_NAME).value,
        "AeroplanNumber":        ws.cell(row=row_num, column=config.COL_AEROPLAN).value,
        "FirstDepartureAirport": ws.cell(row=row_num, column=config.COL_FIRST_DEP).value,
        "OutboundSegments":      ws.cell(row=row_num, column=config.COL_OUTBOUND_SEG).value,
        "ReturnSegments":        ws.cell(row=row_num, column=config.COL_RETURN_SEG).value,
        "MontrealArrivalTime":   ws.cell(row=row_num, column=config.COL_MTL_ARRIVAL).value,
        "MontrealDepartureTime": ws.cell(row=row_num, column=config.COL_MTL_DEPARTURE).value,
        "FlightPassProduct":     ws.cell(row=row_num, column=config.COL_FP_PRODUCT).value,
        "CreditsPerPassenger":   ws.cell(row=row_num, column=config.COL_CREDITS_PER_PAX).value,
        "Cost":                  ws.cell(row=row_num, column=config.COL_COST).value,
        "Type":                  ws.cell(row=row_num, column=config.COL_TYPE).value,
    }


def _remove_from_error_sheet(wb, aeroplan_str: str):
    """Delete the row on the Error sheet whose AeroplanNumber matches."""
    if config.SHEET_ERROR not in wb.sheetnames:
        return
    ws = wb[config.SHEET_ERROR]
    # Error sheet columns: EntryID(1) PNR(2) PassengerName(3) AeroplanNumber(4) Reason(5)
    for row_num in range(ws.max_row, 1, -1):
        cell_val = ws.cell(row=row_num, column=4).value
        if cell_val is None:
            continue
        if str(cell_val).replace(" ", "") == aeroplan_str:
            ws.delete_rows(row_num)


def run_rematch_errors(excel_path: str):
    abs_path = os.path.abspath(excel_path)
    print("Opening workbook...")
    wb        = openpyxl.load_workbook(abs_path)
    wb_lookup = openpyxl.load_workbook(abs_path, data_only=True, read_only=True)

    if config.SHEET_PASSENGER not in wb.sheetnames:
        print("PassengerData sheet not found.")
        wb_lookup.close()
        return

    ws = wb[config.SHEET_PASSENGER]

    error_rows = []
    for row_num in range(2, ws.max_row + 1):
        match_status = ws.cell(row=row_num, column=config.COL_MATCH_STATUS).value
        entry_id     = ws.cell(row=row_num, column=config.COL_ENTRY_ID).value
        if entry_id and str(match_status or "").strip() == "Error":
            error_rows.append(row_num)

    if not error_rows:
        print("No Error rows found — nothing to do.")
        wb_lookup.close()
        return

    print(f"Found {len(error_rows)} Error row(s). Attempting re-match...")
    print()

    resolved   = []
    unresolved = []

    for row_num in error_rows:
        passenger_data = _read_passenger_row_openpyxl(ws, row_num)
        aeroplan       = str(passenger_data.get("AeroplanNumber") or "").strip()
        name           = passenger_data.get("PassengerName") or "(no name)"

        source_sheet, reg_data = lookup_aeroplan(wb_lookup, aeroplan)

        if source_sheet:
            details_name = details_sheet_for(source_sheet)
            details_ws   = ensure_details_sheet(wb, details_name)
            dest_row     = _get_next_row_any(details_ws)
            write_details_row(details_ws, dest_row, passenger_data, reg_data)
            match_label = "Student" if source_sheet == config.SHEET_STUDENT else "Staff"
            ws.cell(row=row_num, column=config.COL_MATCH_STATUS).value = match_label
            _remove_from_error_sheet(wb, aeroplan.replace(" ", ""))
            print(f"  [RESOLVED ] {name:<30} → {match_label}")
            resolved.append(name)
        else:
            reason = "No Aeroplan number" if not aeroplan else "Still not in Student or Staff"
            print(f"  [STILL ERR] {name:<30} — {reason}")
            unresolved.append(name)

    wb_lookup.close()
    wb.save(abs_path)

    print()
    print("-" * 60)
    print(f"  Resolved  : {len(resolved)}")
    print(f"  Remaining : {len(unresolved)}")
    print("-" * 60)
    if unresolved:
        print()
        print("Still unresolved — add these Aeroplan numbers to the Student")
        print("or Staff sheet, then run this again:")
        for n in unresolved:
            print(f"  {n}")
    print()
    print("Done — workbook saved.")
    os.startfile(abs_path)


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else config.EXCEL_FILE
    try:
        run_rematch_errors(path)
    except Exception:
        traceback.print_exc()
    input("\nPress Enter to close...")
