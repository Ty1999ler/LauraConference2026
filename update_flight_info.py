"""
Scan the AC Flight Changes Outlook folder for unread emails with xlsx attachments.
For each attachment, read the updated itinerary rows and overwrite flight columns
in PassengerData and both Plane Details sheets, matching by PNR.

PassengerData columns updated:  E(5) F(6) G(7) H(8) I(9)
Plane Details columns updated:  J(10) K(11) L(12) M(13) N(14)
EmailStatus is never touched.
"""
import os
import sys
import tempfile
import traceback
import openpyxl
import win32com.client
import pythoncom

import config
from outlook_connector import get_outlook_folder

# Plane Details column positions (1-indexed, matches DETAILS_HEADERS in excel_writer.py)
_DET_COL_PNR       = 8   # H
_DET_COL_FIRST_DEP = 10  # J
_DET_COL_OUTBOUND  = 11  # K
_DET_COL_RETURN    = 12  # L
_DET_COL_MTL_ARR   = 13  # M
_DET_COL_MTL_DEP   = 14  # N

# Normalized header substrings → canonical field names
_HEADER_KEYS = {
    "pnr":                   "PNR",
    "firstdepartureairport": "FirstDepartureAirport",
    "outboundsegments":      "OutboundSegments",
    "returnsegments":        "ReturnSegments",
    "montrealarrivaltime":   "MontrealArrivalTime",
    "montrealdeparturetime": "MontrealDepartureTime",
}


def _get_updates_folder(namespace):
    folder = namespace.GetDefaultFolder(6)  # olFolderInbox
    for name in config.FOLDER_PATH_UPDATES[1:]:
        folder = folder.Folders(name)
    return folder


def _read_xlsx_from_attachment(attachment) -> list:
    """
    Save attachment to a temp file, read with openpyxl.
    Returns list of dicts keyed by canonical field names.
    Rows missing a PNR value are skipped.
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        attachment.SaveAsFile(tmp.name)
        wb = openpyxl.load_workbook(tmp.name, data_only=True, read_only=True)
        ws = wb.active

        # Map canonical field name → 0-based column index from row 1 headers
        col_map = {}
        for cell in ws[1]:
            if not cell.value:
                continue
            norm = str(cell.value).replace(" ", "").lower()
            for key, field in _HEADER_KEYS.items():
                if key in norm:
                    col_map[field] = cell.column - 1  # 0-based
                    break

        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not any(raw):
                continue
            record = {}
            for field, idx in col_map.items():
                val = raw[idx] if idx < len(raw) else None
                record[field] = str(val).strip() if val is not None else ""
            if record.get("PNR", "").strip():
                rows.append(record)

        wb.close()
        return rows
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


def _update_passenger_data(ws_com, pnr: str, new_fields: dict, last_row: int) -> list:
    """Overwrite flight columns in PassengerData for all rows matching PNR."""
    updated = []
    for row_num in range(2, last_row + 1):
        cell_pnr = ws_com.Cells(row_num, config.COL_PNR).Value
        if not cell_pnr or str(cell_pnr).strip().upper() != pnr:
            continue
        name = ws_com.Cells(row_num, config.COL_PASSENGER_NAME).Value or f"row {row_num}"
        if "FirstDepartureAirport" in new_fields:
            ws_com.Cells(row_num, config.COL_FIRST_DEP).Value    = new_fields["FirstDepartureAirport"]
        if "OutboundSegments" in new_fields:
            ws_com.Cells(row_num, config.COL_OUTBOUND_SEG).Value = new_fields["OutboundSegments"]
        if "ReturnSegments" in new_fields:
            ws_com.Cells(row_num, config.COL_RETURN_SEG).Value   = new_fields["ReturnSegments"]
        if "MontrealArrivalTime" in new_fields:
            ws_com.Cells(row_num, config.COL_MTL_ARRIVAL).Value  = new_fields["MontrealArrivalTime"]
        if "MontrealDepartureTime" in new_fields:
            ws_com.Cells(row_num, config.COL_MTL_DEPARTURE).Value = new_fields["MontrealDepartureTime"]
        updated.append(str(name).strip())
    return updated


def _update_details_by_pnr(wb_com, sheet_name: str, pnr: str, new_fields: dict) -> list:
    """Overwrite flight columns in a Plane Details sheet for all rows matching PNR."""
    try:
        ws = wb_com.Sheets(sheet_name)
    except Exception:
        return []

    last_row = ws.Cells(ws.Rows.Count, _DET_COL_PNR).End(-4162).Row
    updated = []
    for row_num in range(2, last_row + 1):
        cell_pnr = ws.Cells(row_num, _DET_COL_PNR).Value
        if not cell_pnr or str(cell_pnr).strip().upper() != pnr:
            continue
        name = ws.Cells(row_num, 1).Value or f"row {row_num}"
        if "FirstDepartureAirport" in new_fields:
            ws.Cells(row_num, _DET_COL_FIRST_DEP).Value = new_fields["FirstDepartureAirport"]
        if "OutboundSegments" in new_fields:
            ws.Cells(row_num, _DET_COL_OUTBOUND).Value  = new_fields["OutboundSegments"]
        if "ReturnSegments" in new_fields:
            ws.Cells(row_num, _DET_COL_RETURN).Value    = new_fields["ReturnSegments"]
        if "MontrealArrivalTime" in new_fields:
            ws.Cells(row_num, _DET_COL_MTL_ARR).Value   = new_fields["MontrealArrivalTime"]
        if "MontrealDepartureTime" in new_fields:
            ws.Cells(row_num, _DET_COL_MTL_DEP).Value   = new_fields["MontrealDepartureTime"]
        updated.append(str(name).strip())
    return updated


def run_update_flight_info(excel_path: str):
    pythoncom.CoInitialize()
    abs_path = os.path.abspath(excel_path)

    print("Connecting to Outlook...")
    try:
        outlook = win32com.client.GetActiveObject("Outlook.Application")
    except Exception:
        outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")

    try:
        folder = _get_updates_folder(namespace)
        print(f"Scanning folder: {folder.Name}")
    except Exception as exc:
        print(f"  ERROR: Could not open AC Flight Changes folder — {exc}")
        return

    unread_with_xlsx = []
    for item in folder.Items:
        try:
            if item.Class != 43:
                continue
            if not item.UnRead:
                continue
            for att in item.Attachments:
                if att.FileName.lower().endswith(".xlsx"):
                    unread_with_xlsx.append((item, att))
                    break
        except Exception:
            continue

    if not unread_with_xlsx:
        print("No unread emails with xlsx attachments found.")
        return

    print(f"Found {len(unread_with_xlsx)} unread email(s) with xlsx attachments.")
    print()

    pnr_updates = {}   # PNR (upper) → new_fields dict
    processed_items = []

    for mail_item, attachment in unread_with_xlsx:
        subj = mail_item.Subject or "(no subject)"
        print(f"  Reading: {subj}")
        try:
            rows = _read_xlsx_from_attachment(attachment)
            count = 0
            for row in rows:
                pnr = row.get("PNR", "").strip().upper()
                if not pnr:
                    continue
                fields = {k: v for k, v in row.items() if k != "PNR" and v}
                pnr_updates[pnr] = fields
                count += 1
            processed_items.append(mail_item)
            print(f"    → {count} PNR(s) parsed")
        except Exception as exc:
            print(f"    [ERROR] Could not read attachment — {exc}")

    if not pnr_updates:
        print("No PNR data found in any attachment.")
        return

    print()
    print(f"Updating {len(pnr_updates)} PNR(s) in workbook...")
    print()

    try:
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")

        xl.Visible = False
        wb_com = None
        for w in xl.Workbooks:
            if w.FullName.lower() == abs_path.lower():
                wb_com = w
                break
        if wb_com is None:
            wb_com = xl.Workbooks.Open(abs_path)

        ws_pd    = wb_com.Sheets(config.SHEET_PASSENGER)
        last_pd  = ws_pd.Cells(ws_pd.Rows.Count, config.COL_PNR).End(-4162).Row

        total_pd  = 0
        total_det = 0

        for pnr, new_fields in pnr_updates.items():
            pd_names   = _update_passenger_data(ws_pd, pnr, new_fields, last_pd)
            det_staff  = _update_details_by_pnr(wb_com, config.SHEET_STAFF_DETAILS,   pnr, new_fields)
            det_stu    = _update_details_by_pnr(wb_com, config.SHEET_STUDENT_DETAILS, pnr, new_fields)
            all_names  = list(dict.fromkeys(pd_names + det_staff + det_stu))

            if all_names:
                print(f"  PNR {pnr}: {len(pd_names)} PassengerData, "
                      f"{len(det_staff)+len(det_stu)} Plane Details row(s) updated")
                for n in all_names:
                    print(f"    {n}")
                total_pd  += len(pd_names)
                total_det += len(det_staff) + len(det_stu)
            else:
                print(f"  PNR {pnr}: NOT FOUND in workbook — check PNR is correct")

        wb_com.Save()
        xl.Visible = True

        for mail_item in processed_items:
            try:
                mail_item.UnRead = False
            except Exception:
                pass

        print()
        print(f"Done — {total_pd} PassengerData and {total_det} Plane Details row(s) updated.")

    except Exception as exc:
        print(f"  [ERROR] {exc}")
        traceback.print_exc()


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else config.EXCEL_FILE
    try:
        run_update_flight_info(path)
    except Exception:
        traceback.print_exc()
    input("\nPress Enter to close...")
