import sys
import traceback
import openpyxl

import config
from outlook_connector import get_outlook_folder, get_folder_items
from parse_flight_pass import get_email_type, parse_flight_pass_email
from parse_paid_tickets import parse_paid_email
from match_passenger import lookup_aeroplan, details_sheet_for
from excel_writer import (
    ensure_headers, write_row, format_passenger_sheet,
    get_all_entry_ids, get_next_row, log_debug,
    ensure_details_sheet, ensure_error_sheet,
    write_details_row, write_error_row, _get_next_row_any,
)

REQUIRED_FIELDS = [
    "PNR", "PassengerName", "FirstDepartureAirport",
    "OutboundSegments", "MontrealArrivalTime",
]


def _print_row_summary(row_data: dict, row_num: int):
    name    = row_data.get("PassengerName") or "(no name)"
    pnr     = row_data.get("PNR") or "(no PNR)"
    typ     = row_data.get("Type") or ""
    missing = [f for f in REQUIRED_FIELDS if not row_data.get(f)]
    status  = "OK     " if not missing else "MISSING"
    print(f"  [{status}] Row {row_num:>3} | {pnr:<8} | {name:<30} | {typ}")
    if missing:
        print(f"           Missing: {', '.join(missing)}")


def _read_passenger_row(ws, row_num: int) -> dict:
    return {
        "EntryID":               ws.cell(row=row_num, column=config.COL_ENTRY_ID).value,
        "PNR":                   ws.cell(row=row_num, column=config.COL_PNR).value,
        "PassengerName":         ws.cell(row=row_num, column=config.COL_PASSENGER_NAME).value,
        "AeroplanNumber":        ws.cell(row=row_num, column=config.COL_AEROPLAN).value,
        "FirstDepartureAirport": ws.cell(row=row_num, column=config.COL_FIRST_DEP).value,
        "OutboundSegments":      ws.cell(row=row_num, column=config.COL_OUTBOUND_SEG).value,
        "ReturnSegments":        ws.cell(row=row_num, column=config.COL_RETURN_SEG).value,
        "MontrealArrivalTime":   ws.cell(row=row_num, column=config.COL_MTL_ARRIVAL).value,
        "MontrealDepartureTime": ws.cell(row=row_num, column=config.COL_MTL_DEPARTURE).value,
        "FlightPassProduct":     ws.cell(row=row_num, column=config.COL_FP_PRODUCT).value,
        "CreditsPerPassenger":   ws.cell(row=row_num, column=config.COL_CREDITS_PER_PAX).value,
        "Cost":                  ws.cell(row=row_num, column=config.COL_COST).value,
        "Type":                  ws.cell(row=row_num, column=config.COL_TYPE).value,
    }


def run_everything(excel_path: str):
    print("Opening workbook...")
    wb = openpyxl.load_workbook(excel_path)

    if config.SHEET_PASSENGER not in wb.sheetnames:
        wb.create_sheet(config.SHEET_PASSENGER)
    ws = wb[config.SHEET_PASSENGER]

    if config.SHEET_DEBUG not in wb.sheetnames:
        debug_ws = wb.create_sheet(config.SHEET_DEBUG)
        debug_ws.append(["EntryID", "Subject", "Error", "RowNum"])
    else:
        debug_ws = wb[config.SHEET_DEBUG]

    ensure_headers(ws)
    processed_ids = get_all_entry_ids(ws)
    print(f"Already processed: {len(processed_ids)} email(s)")

    print("Connecting to Outlook...")
    folder = get_outlook_folder(config.FOLDER_PATH)
    items  = get_folder_items(folder)
    print(f"Found {len(items)} email(s) in folder")
    print()

    next_row    = get_next_row(ws)
    added_count = 0
    skip_count  = 0
    error_count = 0

    # ── Step 1: Parse emails → PassengerData ─────────────────────────────
    for mail in items:
        subject  = mail.Subject or ""
        entry_id = mail.EntryID

        if entry_id in processed_ids:
            skip_count += 1
            continue

        email_type = get_email_type(subject)
        if not email_type:
            continue

        print(f"  Processing: {subject[:70]}")

        try:
            body = mail.Body or ""

            if email_type == "flightPass":
                rows = parse_flight_pass_email(body, entry_id)
            else:
                rows = parse_paid_email(body, entry_id, subject)

            if not rows:
                print("  [WARNING] No passengers found in this email")

            for row_data in rows:
                write_row(ws, next_row, row_data)
                _print_row_summary(row_data, next_row)
                next_row    += 1
                added_count += 1

            processed_ids.add(entry_id)

        except Exception as exc:
            err_msg = traceback.format_exc()
            print(f"  [ERROR] {exc}")
            log_debug(debug_ws, entry_id, subject, err_msg, next_row)
            error_count += 1

        print()

    print("-" * 60)
    print(f"  New rows added : {added_count}")
    print(f"  Skipped (dupe) : {skip_count}")
    print(f"  Errors         : {error_count}")
    print("-" * 60)

    # ── Step 2: Match every unmatched PassengerData row by Aeroplan ──────
    print()
    print("Matching passengers to Student / Staff...")

    match_count = 0
    no_match_count = 0

    for row_num in range(2, ws.max_row + 1):
        entry_id = ws.cell(row=row_num, column=config.COL_ENTRY_ID).value
        if not entry_id:
            continue

        # Skip rows already matched
        already = ws.cell(row=row_num, column=config.COL_DETAILS_SHEET).value
        if already:
            continue

        passenger_data = _read_passenger_row(ws, row_num)
        aeroplan       = str(passenger_data.get("AeroplanNumber") or "").strip()

        source_sheet, reg_data = lookup_aeroplan(wb, aeroplan)

        if source_sheet:
            details_name = details_sheet_for(source_sheet)
            details_ws   = ensure_details_sheet(wb, details_name)
            dest_row     = _get_next_row_any(details_ws)
            write_details_row(details_ws, dest_row, passenger_data, reg_data)
            ws.cell(row=row_num, column=config.COL_DETAILS_SHEET).value = details_name
            name = passenger_data.get("PassengerName") or "(no name)"
            print(f"  [MATCHED ] {name:<30} → {details_name}")
            match_count += 1
        else:
            reason   = ("No Aeroplan number" if not aeroplan
                        else "Not found in Student or Staff")
            error_ws = ensure_error_sheet(wb)
            dest_row = _get_next_row_any(error_ws)
            write_error_row(error_ws, dest_row, passenger_data, reason)
            ws.cell(row=row_num, column=config.COL_DETAILS_SHEET).value = "Error"
            name = passenger_data.get("PassengerName") or "(no name)"
            print(f"  [NO MATCH] {name:<30} — {reason}")
            no_match_count += 1

    print()
    print("-" * 60)
    print(f"  Matched   : {match_count}")
    print(f"  No match  : {no_match_count}  (check Error sheet)")
    print("-" * 60)

    print("Applying formatting...")
    format_passenger_sheet(ws)

    wb.save(excel_path)
    print("Done - workbook saved.")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else config.EXCEL_FILE
    run_everything(path)
