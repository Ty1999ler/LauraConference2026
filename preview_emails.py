"""
Batch email preview: opens Outlook forward drafts for unpreviewd passengers.

Order: Staff rows first, then Student rows.
Cap: config.MAX_PREVIEW_EMAILS per run.

Never calls .Send() — always opens as a draft for human review.

On each run, scans Outlook Sent Items (from config.SENT_SCAN_CUTOFF onwards)
first: any Previewed row whose original email's ConversationID matches a sent
forward is automatically marked "Sent".

Uses openpyxl read-only for all reads, win32com for cell updates and save
so openpyxl never writes the file (prevents xlsm data corruption).

Validation errors (missing email, original email not in Outlook) are flagged
in the console and written to EmailStatus so they show up in the spreadsheet.
"""
import os
import re
import sys
import traceback
import openpyxl
import win32com.client
import pythoncom

import config

# Column positions in the Plane Details sheets (1-indexed, matches DETAILS_HEADERS)
_DETAILS_COL_PREFERRED_NAME = 2   # B — Preferred Name
_DETAILS_COL_EMAIL          = 3   # C — Email
_DETAILS_COL_AEROPLAN       = 7   # G — AeroplanNumber
_DETAILS_COL_EMAIL_STATUS   = 18  # R — EmailStatus


def _get_outlook():
    pythoncom.CoInitialize()
    try:
        return win32com.client.GetActiveObject("Outlook.Application")
    except Exception:
        return win32com.client.Dispatch("Outlook.Application")


def _open_forward_draft(namespace, entry_id: str, preferred_name: str, to_address: str):
    """Open a forward draft addressed to to_address. Never calls .Send()."""
    item   = namespace.GetItemFromID(entry_id)
    fwd    = item.Forward()
    fwd.To = to_address

    name_part = preferred_name if preferred_name else ""
    intro = (
        f"<p>Hi {name_part},</p>"
        f"<p>I'm very excited to welcome you to the inaugural Alumo Summit!</p>"
        f"<p>Please find your travel booking below!</p>"
        f"<p>I'll be sharing additional information about the conference in June so stay tuned! "
        f"This will include:</p>"
        f"<ul><li>Summit agenda</li><li>Accommodation details</li><li>Shuttle schedule</li>"
        f"<li>Meal options</li><li>App details</li><li>And much more!!</li></ul>"
        f"<p>In the meantime, if you have any questions, please don't hesitate to reach out.</p>"
        f"<p>The Alumo team is excited to welcome you to Tremblant this July!</p>"
        f"<p>Looking forward to seeing you soon,</p>"
    )

    html  = fwd.HTMLBody
    match = re.search(r'<body[^>]*>', html, re.IGNORECASE)
    if match:
        pos          = match.end()
        fwd.HTMLBody = html[:pos] + intro + html[pos:]
    else:
        fwd.HTMLBody = intro + html

    fwd.Subject = "Alumo Summit – Travel Booking"
    fwd.Display()  # preview only — NEVER .Send()


def _build_details_maps(wb_ro) -> tuple:
    """
    Returns ({aeroplan_str: preferred_name}, {aeroplan_str: email})
    from both Plane Details sheets.
    """
    names  = {}
    emails = {}
    for sheet_name in [config.SHEET_STUDENT_DETAILS, config.SHEET_STAFF_DETAILS]:
        if sheet_name not in wb_ro.sheetnames:
            continue
        ws = wb_ro[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < _DETAILS_COL_AEROPLAN:
                continue
            aeroplan = row[_DETAILS_COL_AEROPLAN - 1]
            if not aeroplan:
                continue
            if isinstance(aeroplan, float):
                aeroplan = int(aeroplan)
            key           = str(aeroplan).replace(' ', '')
            pref          = row[_DETAILS_COL_PREFERRED_NAME - 1]
            email         = row[_DETAILS_COL_EMAIL - 1]
            names[key]    = str(pref).strip()  if pref  else ''
            emails[key]   = str(email).strip() if email else ''
    return names, emails


def _find_sent_entry_ids(namespace, previewed_rows: list) -> set:
    """
    Returns entry_ids of previewed rows whose Alumo Summit forward has been sent.
    Matches by finding sent items with 'alumo summit' in the subject addressed to
    the passenger — ConversationID is not used because changing the subject on a
    forward creates a new conversation in Outlook.
    """
    if not previewed_rows:
        return set()

    print(f"  Scanning Sent Items (from {config.SENT_SCAN_CUTOFF}) for completed forwards...")

    sent_folder = namespace.GetDefaultFolder(5)  # olFolderSentMail
    restricted  = sent_folder.Items.Restrict(
        f"[SentOn] >= '{config.SENT_SCAN_CUTOFF}'"
    )

    sent_to_addresses = set()
    for item in restricted:
        try:
            if (item.Subject or '').lower() != 'alumo summit - travel booking':
                continue
            for recip in item.Recipients:
                try:
                    addr = (recip.Address or '').lower().strip()
                    if addr:
                        sent_to_addresses.add(addr)
                except Exception:
                    continue
        except Exception:
            continue

    if not sent_to_addresses:
        print("  No matching sent forwards found.")
        return set()

    matched = set()
    for entry_id, _name, _aeroplan, _match, to_email in previewed_rows:
        if to_email.lower() in sent_to_addresses:
            matched.add(entry_id)

    print(f"  Found {len(matched)} sent forward(s).")
    return matched


def _update_details_sheet(wb_com, sheet_name: str, aeroplan_str: str, status: str):
    """Set EmailStatus (col R) in the Plane Details sheet for the matching Aeroplan row."""
    try:
        ws = wb_com.Sheets(sheet_name)
    except Exception:
        return
    last_row = ws.Cells(ws.Rows.Count, _DETAILS_COL_AEROPLAN).End(-4162).Row  # xlUp
    for row_num in range(2, last_row + 1):
        cell_val = ws.Cells(row_num, _DETAILS_COL_AEROPLAN).Value
        if not cell_val:
            continue
        cell_str = str(int(cell_val)) if isinstance(cell_val, float) else str(cell_val).replace(' ', '')
        if cell_str == aeroplan_str:
            ws.Cells(row_num, _DETAILS_COL_EMAIL_STATUS).Value = status
            return


def run_preview(excel_path: str):
    pythoncom.CoInitialize()

    abs_path = os.path.abspath(excel_path)

    # ── Read workbook (read-only — never writes) ──────────────────────────
    print("Reading workbook...")
    wb_ro = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)

    if config.SHEET_PASSENGER not in wb_ro.sheetnames:
        print("PassengerData sheet not found — nothing to do.")
        wb_ro.close()
        return

    preferred_name_map, email_map = _build_details_maps(wb_ro)

    ws_ro          = wb_ro[config.SHEET_PASSENGER]
    staff_rows     = []   # valid, unpreviewd, Staff
    student_rows   = []   # valid, unpreviewd, Student
    previewed_rows = []   # EmailStatus == "Previewed" → check if now sent
    # validation_errors: (entry_id, aeroplan_str, match_str, error_msg)
    validation_errors = []

    print()
    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        if len(row) < config.COL_MATCH_STATUS:
            continue

        entry_id     = row[config.COL_ENTRY_ID      - 1]
        email_status = row[config.COL_EMAIL_STATUS   - 1]
        match_status = row[config.COL_MATCH_STATUS   - 1]
        name         = str(row[config.COL_PASSENGER_NAME - 1] or "")
        aeroplan     = row[config.COL_AEROPLAN       - 1]

        if not entry_id:
            continue

        # Skip rows already fully processed
        if email_status in ("Sent",) or (email_status and str(email_status).startswith("Error:")):
            continue

        aeroplan_str   = str(aeroplan).replace(' ', '') if aeroplan else ''
        match_str      = str(match_status or '')

        # Only process Staff and Student rows
        if match_str not in ("Staff", "Student"):
            continue

        preferred_name = preferred_name_map.get(aeroplan_str, '') or name
        to_email       = email_map.get(aeroplan_str, '')

        # ── Validate required fields ──────────────────────────────────────
        if not to_email:
            msg = "No email address in Plane Details"
            print(f"  [ERR] {name or entry_id} — {msg}")
            validation_errors.append((str(entry_id), aeroplan_str, match_str, msg))
            continue

        if not preferred_name:
            print(f"  [WARN] {name or entry_id} — No preferred name, using passenger name")

        record = (str(entry_id), preferred_name, aeroplan_str, match_str, to_email)

        if email_status == "Previewed":
            previewed_rows.append(record)
        elif not email_status:
            if match_str == "Staff":
                staff_rows.append(record)
            else:
                student_rows.append(record)

    wb_ro.close()

    if validation_errors:
        print()
        print(f"  {len(validation_errors)} passenger(s) need attention before they can be emailed.")

    # ── Connect to Outlook ────────────────────────────────────────────────
    print()
    outlook   = _get_outlook()
    namespace = outlook.GetNamespace("MAPI")

    # ── Step 1: Check Sent Items for previously previewed rows ────────────
    newly_sent_ids = _find_sent_entry_ids(namespace, previewed_rows)
    newly_sent_map = {r[0]: r for r in previewed_rows if r[0] in newly_sent_ids}

    if newly_sent_ids:
        print(f"  Marking {len(newly_sent_ids)} row(s) as Sent.")
    print()

    # ── Step 2: Open forward drafts for unpreviewd rows ───────────────────
    to_preview    = staff_rows + student_rows
    new_previewed = []   # (entry_id, aeroplan_str, match_status)
    new_errored   = []   # (entry_id, aeroplan_str, match_status, error_msg)

    if not to_preview:
        if not newly_sent_ids and not validation_errors:
            print("Nothing to do — all passengers already previewed or sent.")
    else:
        total     = len(to_preview)
        cap       = config.MAX_PREVIEW_EMAILS
        batch     = to_preview[:cap]
        remaining = total - len(batch)

        print(f"Found {total} unpreviewd passenger(s) — opening {len(batch)} draft(s)...")
        print(f"  ({len(staff_rows)} Staff, {len(student_rows)} Student pending)")
        print()

        for entry_id, preferred_name, aeroplan_str, match_status, to_email in batch:
            try:
                _open_forward_draft(namespace, entry_id, preferred_name, to_email)
                new_previewed.append((entry_id, aeroplan_str, match_status))
                print(f"  [OK ] {preferred_name or entry_id[:12]} → {to_email}")
            except Exception as exc:
                msg = str(exc)
                new_errored.append((entry_id, aeroplan_str, match_status, msg))
                print(f"  [ERR] {preferred_name or entry_id[:12]} — {msg}")

        if remaining:
            print()
            print(f"{remaining} more unpreviewd — run again for next batch of {cap}.")

    # ── Step 3: Write all status updates via win32com ─────────────────────
    all_updates = newly_sent_ids or new_previewed or new_errored or validation_errors
    if not all_updates:
        return

    print()
    print("Saving updates via Excel...")
    try:
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")

        xl.Visible = False
        wb_com = None
        for w in xl.Workbooks:
            if w.FullName.lower() == abs_path.lower():
                wb_com = w
                break
        if wb_com is None:
            wb_com = xl.Workbooks.Open(abs_path)

        ws_com   = wb_com.Sheets(config.SHEET_PASSENGER)
        last_row = ws_com.Cells(ws_com.Rows.Count, config.COL_ENTRY_ID).End(-4162).Row

        sent_map      = newly_sent_map
        previewed_map = {eid: (ap, ms) for eid, ap, ms in new_previewed}
        errored_map   = {eid: (ap, ms, msg) for eid, ap, ms, msg in new_errored}
        val_err_map   = {eid: (ap, ms, msg) for eid, ap, ms, msg in validation_errors}

        for row_num in range(2, last_row + 1):
            eid = ws_com.Cells(row_num, config.COL_ENTRY_ID).Value
            if not eid:
                continue
            eid_str = str(eid)

            if eid_str in sent_map:
                _, _, aeroplan_str, match_status = sent_map[eid_str]
                ws_com.Cells(row_num, config.COL_EMAIL_STATUS).Value = "Sent"
                details = (config.SHEET_STAFF_DETAILS if match_status == "Staff"
                           else config.SHEET_STUDENT_DETAILS)
                if aeroplan_str:
                    _update_details_sheet(wb_com, details, aeroplan_str, "Sent")

            elif eid_str in previewed_map:
                aeroplan_str, match_status = previewed_map[eid_str]
                ws_com.Cells(row_num, config.COL_EMAIL_STATUS).Value = "Previewed"
                details = (config.SHEET_STAFF_DETAILS if match_status == "Staff"
                           else config.SHEET_STUDENT_DETAILS)
                if aeroplan_str:
                    _update_details_sheet(wb_com, details, aeroplan_str, "Previewed")

            elif eid_str in errored_map:
                _, _, msg = errored_map[eid_str]
                ws_com.Cells(row_num, config.COL_EMAIL_STATUS).Value = f"Error: {msg}"

            elif eid_str in val_err_map:
                _, _, msg = val_err_map[eid_str]
                ws_com.Cells(row_num, config.COL_EMAIL_STATUS).Value = f"Error: {msg}"

        wb_com.Save()
        xl.Visible = True
        print("Saved.")

    except Exception as exc:
        print(f"  [WARNING] Could not save updates — {exc}")
        print("  Open the workbook manually and re-run to retry.")

    print()
    summary = []
    if newly_sent_ids:
        summary.append(f"{len(newly_sent_ids)} marked Sent")
    if new_previewed:
        summary.append(f"{len(new_previewed)} draft(s) opened")
    if new_errored:
        summary.append(f"{len(new_errored)} draft error(s)")
    if validation_errors:
        summary.append(f"{len(validation_errors)} need attention (check EmailStatus)")
    print("Done" + (" — " + ", ".join(summary) if summary else "") + ".")


def run_check_forwards(excel_path: str):
    """Scan Sent Items only — mark Previewed rows as Sent. No new drafts opened."""
    pythoncom.CoInitialize()

    abs_path = os.path.abspath(excel_path)

    print("Reading workbook...")
    wb_ro = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)

    if config.SHEET_PASSENGER not in wb_ro.sheetnames:
        print("PassengerData sheet not found — nothing to do.")
        wb_ro.close()
        return

    preferred_name_map, email_map = _build_details_maps(wb_ro)

    ws_ro          = wb_ro[config.SHEET_PASSENGER]
    previewed_rows = []

    for row in ws_ro.iter_rows(min_row=2, values_only=True):
        if len(row) < config.COL_MATCH_STATUS:
            continue
        entry_id     = row[config.COL_ENTRY_ID     - 1]
        email_status = row[config.COL_EMAIL_STATUS  - 1]
        match_status = row[config.COL_MATCH_STATUS  - 1]
        aeroplan     = row[config.COL_AEROPLAN      - 1]
        name         = str(row[config.COL_PASSENGER_NAME - 1] or "")

        if not entry_id or email_status not in ("Previewed", None, ""):
            continue
        if str(match_status or '') not in ("Staff", "Student"):
            continue

        if aeroplan:
            ap = int(aeroplan) if isinstance(aeroplan, float) else aeroplan
            aeroplan_str = str(ap).replace(' ', '')
        else:
            aeroplan_str = ''

        to_email = email_map.get(aeroplan_str, '')
        if not to_email:
            continue

        preferred_name = preferred_name_map.get(aeroplan_str, '') or name
        previewed_rows.append((str(entry_id), preferred_name, aeroplan_str,
                               str(match_status), to_email))

    wb_ro.close()

    if not previewed_rows:
        print("No Previewed rows to check.")
        return

    print(f"Found {len(previewed_rows)} Previewed row(s) to check...")

    outlook        = _get_outlook()
    namespace      = outlook.GetNamespace("MAPI")
    newly_sent_ids = _find_sent_entry_ids(namespace, previewed_rows)

    if not newly_sent_ids:
        print("No newly sent forwards found.")
        return

    sent_map = {r[0]: r for r in previewed_rows if r[0] in newly_sent_ids}

    print(f"Marking {len(newly_sent_ids)} row(s) as Sent...")
    try:
        try:
            xl = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            xl = win32com.client.Dispatch("Excel.Application")

        xl.Visible = False
        wb_com = None
        for w in xl.Workbooks:
            if w.FullName.lower() == abs_path.lower():
                wb_com = w
                break
        if wb_com is None:
            wb_com = xl.Workbooks.Open(abs_path)

        ws_com   = wb_com.Sheets(config.SHEET_PASSENGER)
        last_row = ws_com.Cells(ws_com.Rows.Count, config.COL_ENTRY_ID).End(-4162).Row

        for row_num in range(2, last_row + 1):
            eid = ws_com.Cells(row_num, config.COL_ENTRY_ID).Value
            if not eid or str(eid) not in newly_sent_ids:
                continue
            ws_com.Cells(row_num, config.COL_EMAIL_STATUS).Value = "Sent"
            _, _, aeroplan_str, match_status, _ = sent_map[str(eid)]
            details = (config.SHEET_STAFF_DETAILS if match_status == "Staff"
                       else config.SHEET_STUDENT_DETAILS)
            if aeroplan_str:
                _update_details_sheet(wb_com, details, aeroplan_str, "Sent")

        wb_com.Save()
        xl.Visible = True
        print(f"Done — {len(newly_sent_ids)} row(s) marked Sent.")

    except Exception as exc:
        print(f"  [WARNING] Could not save updates — {exc}")


if __name__ == "__main__":
    _args     = sys.argv[1:]
    _paths    = [a for a in _args if not a.startswith("--")]
    path      = _paths[0] if _paths else config.EXCEL_FILE
    try:
        if "--check" in _args:
            run_check_forwards(path)
        else:
            run_preview(path)
    except Exception:
        traceback.print_exc()
    input("\nPress Enter to close...")
